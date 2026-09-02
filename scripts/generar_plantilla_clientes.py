from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "plantillas"
OUTPUT_PATH = OUTPUT_DIR / "Plantilla_Cargue_Masivo_Clientes_Planetour.xlsx"
MAX_DATA_ROW = 5000

NAVY = "132238"
NAVY_LIGHT = "1E3A5F"
INDIGO = "4F46E5"
SKY = "0EA5E9"
ORANGE = "F97316"
RED = "C2413B"
GREEN = "15803D"
PALE_GREEN = "DCFCE7"
PALE_ORANGE = "FFEDD5"
PALE_RED = "FEE2E2"
PALE_BLUE = "E0F2FE"
PALE_GRAY = "F1F5F9"
WHITE = "FFFFFF"
TEXT = "172033"
MUTED = "475569"
BORDER_COLOR = "CBD5E1"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


FIELDS = [
    {
        "key": "nombre_cliente",
        "label": "Razón social / nombre comercial",
        "required": "API",
        "type": "Texto",
        "limit": "2 a 200 caracteres",
        "api": "name",
        "default": "Sin valor predeterminado",
        "example": "Agencia Horizonte Demo S.A.S.",
        "width": 34,
        "description": "Nombre legal o comercial con el que se identificará al cliente.",
    },
    {
        "key": "tipo_cliente",
        "label": "Tipo de cliente",
        "required": "API",
        "type": "Catálogo",
        "limit": "AGENCY / GOVERNMENT / CORPORATE",
        "api": "type",
        "default": "Sin valor predeterminado",
        "example": "AGENCY",
        "width": 20,
        "description": "Clasificación principal del cliente. Usa el código exacto del catálogo.",
    },
    {
        "key": "nit",
        "label": "NIT / RUT",
        "required": "API",
        "type": "Texto",
        "limit": "3 a 50 caracteres; único",
        "api": "nit",
        "default": "Sin valor predeterminado",
        "example": "900123456-7",
        "width": 20,
        "description": "Identificación tributaria. Mantén un único formato para evitar duplicados visuales.",
    },
    {
        "key": "codigo_iata",
        "label": "Código IATA / Pseudo Office",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo 50 caracteres",
        "api": "iataCode",
        "default": "Vacío",
        "example": "76-54321",
        "width": 22,
        "description": "Código IATA o identificador de oficina principal, si aplica.",
    },
    {
        "key": "categoria",
        "label": "Categoría / tier",
        "required": "No",
        "type": "Catálogo",
        "limit": "GOLD / SILVER / BRONZE / ESTATAL",
        "api": "tier",
        "default": "GOLD",
        "example": "GOLD",
        "width": 18,
        "description": "Segmentación comercial. Si queda vacía la aplicación usa GOLD.",
    },
    {
        "key": "estado",
        "label": "Estado inicial",
        "required": "No",
        "type": "Catálogo",
        "limit": "ACTIVE / WARNING / BLOCKED / INACTIVE",
        "api": "status",
        "default": "ACTIVE",
        "example": "ACTIVE",
        "width": 18,
        "description": "Para altas nuevas usa normalmente ACTIVE o INACTIVE; Karing puede recalcular mora y bloqueo.",
    },
    {
        "key": "ciudad",
        "label": "Ciudad principal",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo 100 caracteres",
        "api": "city",
        "default": "Vacío",
        "example": "Bogotá D.C.",
        "width": 22,
        "description": "Ciudad principal de operación o domicilio.",
    },
    {
        "key": "direccion",
        "label": "Dirección",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo 255 caracteres",
        "api": "address",
        "default": "Vacío",
        "example": "Carrera 15 # 93-47",
        "width": 30,
        "description": "Dirección física principal del cliente.",
    },
    {
        "key": "telefono_principal",
        "label": "Teléfono principal",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo 50 caracteres",
        "api": "phone",
        "default": "Vacío",
        "example": "+57 601 555 0101",
        "width": 22,
        "description": "Teléfono general de la empresa. No conviertas esta columna a número.",
    },
    {
        "key": "cupo_credito_cop",
        "label": "Cupo de crédito COP",
        "required": "No",
        "type": "Número",
        "limit": "0 a 999.999.999.999; máximo 2 decimales",
        "api": "creditLimit",
        "default": "0",
        "example": 50000000,
        "width": 22,
        "description": "Cupo aprobado en pesos colombianos, sin símbolos ni separadores escritos manualmente.",
    },
    {
        "key": "dueno_nombre",
        "label": "Nombre del dueño / representante",
        "required": "Operativo",
        "type": "Texto",
        "limit": "Máximo recomendado 150 caracteres",
        "api": "owner.name",
        "default": "Sin valor predeterminado",
        "example": "Carlos Eduardo Mendoza",
        "width": 31,
        "description": "Nombre completo del dueño o representante legal; obligatorio en el formulario de la app.",
    },
    {
        "key": "dueno_documento",
        "label": "Documento del dueño",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 50 caracteres",
        "api": "owner.document",
        "default": "Vacío",
        "example": "CC 79845120",
        "width": 22,
        "description": "Documento del dueño o representante. No se muestra en la consulta pública.",
    },
    {
        "key": "dueno_telefono",
        "label": "Teléfono del dueño",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 50 caracteres",
        "api": "owner.phone",
        "default": "Vacío",
        "example": "+57 310 892 4410",
        "width": 22,
        "description": "Teléfono directo del dueño o representante.",
    },
    {
        "key": "dueno_correo",
        "label": "Correo del dueño",
        "required": "No",
        "type": "Correo",
        "limit": "Máximo recomendado 254 caracteres",
        "api": "owner.email",
        "default": "Vacío",
        "example": "carlos.mendoza@example.com",
        "width": 34,
        "description": "Correo directo del dueño o representante.",
    },
    {
        "key": "dueno_notas",
        "label": "Notas internas del dueño",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 1.000 caracteres",
        "api": "owner.notes",
        "default": "Vacío",
        "example": "Prefiere contacto por WhatsApp.",
        "width": 34,
        "description": "Notas internas. No se muestran en la consulta pública.",
    },
    {
        "key": "pagos_nombre",
        "label": "Nombre del encargado de pagos",
        "required": "Operativo",
        "type": "Texto",
        "limit": "Máximo recomendado 150 caracteres",
        "api": "accountsPayable.name",
        "default": "Sin valor predeterminado",
        "example": "Laura Ximena Torres",
        "width": 31,
        "description": "Responsable de tesorería, facturación o cartera; obligatorio en el formulario.",
    },
    {
        "key": "pagos_cargo",
        "label": "Cargo del encargado de pagos",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 100 caracteres",
        "api": "accountsPayable.role",
        "default": "Tesorero / Cartera",
        "example": "Tesorera / Cartera",
        "width": 27,
        "description": "Cargo o rol del contacto de pagos.",
    },
    {
        "key": "pagos_telefono",
        "label": "Teléfono de pagos",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 50 caracteres",
        "api": "accountsPayable.phone",
        "default": "Vacío",
        "example": "+57 601 745 9008",
        "width": 22,
        "description": "Teléfono o celular del contacto de pagos.",
    },
    {
        "key": "pagos_whatsapp",
        "label": "WhatsApp de pagos",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 50 caracteres",
        "api": "accountsPayable.whatsapp",
        "default": "Vacío",
        "example": "+57 315 443 8901",
        "width": 22,
        "description": "Número directo de WhatsApp para cartera o facturación.",
    },
    {
        "key": "pagos_correo",
        "label": "Correo de facturación",
        "required": "No",
        "type": "Correo",
        "limit": "Máximo recomendado 254 caracteres",
        "api": "accountsPayable.email",
        "default": "Vacío",
        "example": "pagos@example.com",
        "width": 34,
        "description": "Correo de facturación, tesorería o cartera.",
    },
    {
        "key": "pagos_dias",
        "label": "Día habitual de pagos",
        "required": "No",
        "type": "Catálogo",
        "limit": "Día de la semana",
        "api": "accountsPayable.paymentDays",
        "default": "Viernes",
        "example": "Viernes",
        "width": 22,
        "description": "Día habitual indicado por el cliente para procesar pagos.",
    },
    {
        "key": "operativo_nombre",
        "label": "Nombre del contacto operativo",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 150 caracteres",
        "api": "operationalCounter.name",
        "default": "Vacío",
        "example": "Andrea Ruiz",
        "width": 30,
        "description": "Counter o contacto operativo principal.",
    },
    {
        "key": "operativo_cargo",
        "label": "Cargo del contacto operativo",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 100 caracteres",
        "api": "operationalCounter.role",
        "default": "Counter Principal",
        "example": "Counter Principal",
        "width": 28,
        "description": "Cargo o función del contacto operativo.",
    },
    {
        "key": "operativo_telefono",
        "label": "Teléfono del contacto operativo",
        "required": "No",
        "type": "Texto",
        "limit": "Máximo recomendado 50 caracteres",
        "api": "operationalCounter.phone",
        "default": "Vacío",
        "example": "+57 300 555 2244",
        "width": 27,
        "description": "Teléfono directo del contacto operativo.",
    },
    {
        "key": "operativo_correo",
        "label": "Correo del contacto operativo",
        "required": "No",
        "type": "Correo",
        "limit": "Máximo recomendado 254 caracteres",
        "api": "operationalCounter.email",
        "default": "Vacío",
        "example": "operaciones@example.com",
        "width": 34,
        "description": "Correo directo del contacto operativo.",
    },
]


def style_title(cell, size=20):
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=size, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_section(cell, color=INDIGO):
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name="Aptos", size=12, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")


def add_table(ws, ref, name, style="TableStyleMedium2"):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_named_range(workbook, name, reference):
    workbook.defined_names.add(DefinedName(name, attr_text=reference))


def add_list_validation(ws, formula, cell_range, prompt):
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.errorTitle = "Valor no permitido"
    validation.error = "Selecciona un valor del catálogo disponible."
    validation.promptTitle = "Catálogo Planetour"
    validation.prompt = prompt
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_text_validation(ws, column, maximum, minimum=0, required=False):
    if minimum:
        validation = DataValidation(
            type="textLength",
            operator="between",
            formula1=str(minimum),
            formula2=str(maximum),
            allow_blank=not required,
        )
        rule = f"entre {minimum} y {maximum} caracteres"
    else:
        validation = DataValidation(
            type="textLength",
            operator="lessThanOrEqual",
            formula1=str(maximum),
            allow_blank=True,
        )
        rule = f"máximo {maximum} caracteres"
    validation.errorTitle = "Longitud no permitida"
    validation.error = f"Este campo admite {rule}."
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    validation.add(f"{column}2:{column}{MAX_DATA_ROW}")


def configure_client_sheet(ws):
    ws.sheet_properties.tabColor = INDIGO
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 70
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{MAX_DATA_ROW}"
    ws.row_dimensions[1].height = 48
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.center.text = "Planetour CRM - Cargue masivo de clientes"
    ws.oddFooter.right.text = "Página &P de &N"

    for index, field in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=index, value=field["key"])
        if field["required"] == "API":
            fill_color = RED
            obligation = "OBLIGATORIO PARA LA API"
        elif field["required"] == "Operativo":
            fill_color = ORANGE
            obligation = "OBLIGATORIO OPERATIVO / FORMULARIO"
        else:
            fill_color = NAVY_LIGHT
            obligation = "OPCIONAL"
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.comment = Comment(
            f"{obligation}\n{field['label']}\n{field['description']}\n"
            f"Tipo/límite: {field['type']} — {field['limit']}\n"
            f"Predeterminado: {field['default']}\nMapeo API: {field['api']}",
            "Planetour CRM",
        )
        ws.column_dimensions[get_column_letter(index)].width = field["width"]

    # Texto para conservar ceros, signos, prefijos y formatos de identificación.
    text_keys = {
        "nombre_cliente",
        "nit",
        "codigo_iata",
        "ciudad",
        "direccion",
        "telefono_principal",
        "dueno_nombre",
        "dueno_documento",
        "dueno_telefono",
        "dueno_correo",
        "dueno_notas",
        "pagos_nombre",
        "pagos_cargo",
        "pagos_telefono",
        "pagos_whatsapp",
        "pagos_correo",
        "operativo_nombre",
        "operativo_cargo",
        "operativo_telefono",
        "operativo_correo",
    }
    field_columns = {field["key"]: get_column_letter(index) for index, field in enumerate(FIELDS, 1)}
    for key in text_keys:
        column = field_columns[key]
        for row in range(2, MAX_DATA_ROW + 1):
            ws[f"{column}{row}"].number_format = "@"

    credit_column = field_columns["cupo_credito_cop"]
    for row in range(2, MAX_DATA_ROW + 1):
        ws[f"{credit_column}{row}"].number_format = '#,##0.00;[Red]-#,##0.00'

    add_list_validation(
        ws,
        "=TiposCliente",
        f"{field_columns['tipo_cliente']}2:{field_columns['tipo_cliente']}{MAX_DATA_ROW}",
        "AGENCY, GOVERNMENT o CORPORATE.",
    )
    add_list_validation(
        ws,
        "=CategoriasCliente",
        f"{field_columns['categoria']}2:{field_columns['categoria']}{MAX_DATA_ROW}",
        "GOLD, SILVER, BRONZE o ESTATAL. Vacío equivale a GOLD.",
    )
    add_list_validation(
        ws,
        "=EstadosCliente",
        f"{field_columns['estado']}2:{field_columns['estado']}{MAX_DATA_ROW}",
        "Para altas nuevas usa normalmente ACTIVE o INACTIVE.",
    )
    add_list_validation(
        ws,
        "=DiasPago",
        f"{field_columns['pagos_dias']}2:{field_columns['pagos_dias']}{MAX_DATA_ROW}",
        "Selecciona el día habitual de pagos.",
    )

    numeric_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="999999999999",
        allow_blank=True,
    )
    numeric_validation.errorTitle = "Cupo no válido"
    numeric_validation.error = "Ingresa un número entre 0 y 999.999.999.999."
    numeric_validation.promptTitle = "Valor en COP"
    numeric_validation.prompt = "Escribe solo el número; no agregues $ ni separadores manuales."
    numeric_validation.showErrorMessage = True
    numeric_validation.showInputMessage = True
    ws.add_data_validation(numeric_validation)
    numeric_validation.add(f"{credit_column}2:{credit_column}{MAX_DATA_ROW}")

    validation_limits = {
        "nombre_cliente": (200, 2, True),
        "nit": (50, 3, True),
        "codigo_iata": (50, 0, False),
        "ciudad": (100, 0, False),
        "direccion": (255, 0, False),
        "telefono_principal": (50, 0, False),
        "dueno_nombre": (150, 1, True),
        "dueno_documento": (50, 0, False),
        "dueno_telefono": (50, 0, False),
        "dueno_correo": (254, 0, False),
        "dueno_notas": (1000, 0, False),
        "pagos_nombre": (150, 1, True),
        "pagos_cargo": (100, 0, False),
        "pagos_telefono": (50, 0, False),
        "pagos_whatsapp": (50, 0, False),
        "pagos_correo": (254, 0, False),
        "operativo_nombre": (150, 0, False),
        "operativo_cargo": (100, 0, False),
        "operativo_telefono": (50, 0, False),
        "operativo_correo": (254, 0, False),
    }
    for key, (maximum, minimum, required) in validation_limits.items():
        add_text_validation(ws, field_columns[key], maximum, minimum, required)

    last_column = get_column_letter(len(FIELDS))
    for key in ("nombre_cliente", "tipo_cliente", "nit"):
        column = field_columns[key]
        ws.conditional_formatting.add(
            f"{column}2:{column}{MAX_DATA_ROW}",
            FormulaRule(
                formula=[f'AND(COUNTA($A2:${last_column}2)>0,{column}2="")'],
                fill=PatternFill("solid", fgColor=PALE_RED),
            ),
        )
    for key in ("dueno_nombre", "pagos_nombre"):
        column = field_columns[key]
        ws.conditional_formatting.add(
            f"{column}2:{column}{MAX_DATA_ROW}",
            FormulaRule(
                formula=[f'AND(COUNTA($A2:${last_column}2)>0,{column}2="")'],
                fill=PatternFill("solid", fgColor=PALE_ORANGE),
            ),
        )

    nit_column = field_columns["nit"]
    ws.conditional_formatting.add(
        f"{nit_column}2:{nit_column}{MAX_DATA_ROW}",
        FormulaRule(
            formula=[f'AND({nit_column}2<>"",COUNTIF(${nit_column}$2:${nit_column}${MAX_DATA_ROW},{nit_column}2)>1)'],
            fill=PatternFill("solid", fgColor="FDE68A"),
        ),
    )

    for key in ("dueno_correo", "pagos_correo", "operativo_correo"):
        column = field_columns[key]
        ws.conditional_formatting.add(
            f"{column}2:{column}{MAX_DATA_ROW}",
            FormulaRule(
                formula=[f'AND({column}2<>"",ISERROR(SEARCH("@",{column}2)))'],
                fill=PatternFill("solid", fgColor="FEF3C7"),
            ),
        )


def build_catalogs(ws, workbook):
    ws.sheet_properties.tabColor = GREEN
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    catalogs = [
        (
            1,
            "tipo_cliente",
            [
                ("AGENCY", "Agencia de viajes / minorista"),
                ("GOVERNMENT", "Entidad pública"),
                ("CORPORATE", "Cliente corporativo directo"),
            ],
            "TiposCliente",
        ),
        (
            4,
            "categoria",
            [
                ("GOLD", "Alta emisión"),
                ("SILVER", "Estándar"),
                ("BRONZE", "Ocasional"),
                ("ESTATAL", "Convenio público"),
            ],
            "CategoriasCliente",
        ),
        (
            7,
            "estado",
            [
                ("ACTIVE", "Activo"),
                ("WARNING", "Advertencia de cartera"),
                ("BLOCKED", "Bloqueado"),
                ("INACTIVE", "Inactivo"),
            ],
            "EstadosCliente",
        ),
        (
            10,
            "pagos_dias",
            [
                ("Lunes", "Lunes"),
                ("Martes", "Martes"),
                ("Miércoles", "Miércoles"),
                ("Jueves", "Jueves"),
                ("Viernes", "Viernes"),
                ("Sábado", "Sábado"),
                ("Domingo", "Domingo"),
            ],
            "DiasPago",
        ),
    ]

    for start_column, title, values, range_name in catalogs:
        code_cell = ws.cell(1, start_column, title)
        description_cell = ws.cell(1, start_column + 1, "Descripción")
        for cell in (code_cell, description_cell):
            style_section(cell, NAVY_LIGHT)
            cell.border = THIN_BORDER
        for row_index, (code, description) in enumerate(values, start=2):
            ws.cell(row_index, start_column, code)
            ws.cell(row_index, start_column + 1, description)
            for column_index in (start_column, start_column + 1):
                ws.cell(row_index, column_index).border = THIN_BORDER
                ws.cell(row_index, column_index).fill = PatternFill(
                    "solid", fgColor=WHITE if row_index % 2 else PALE_GRAY
                )
        letter = get_column_letter(start_column)
        add_named_range(workbook, range_name, f"'Catalogos'!${letter}$2:${letter}${len(values) + 1}")

    for column in ("A", "D", "G", "J"):
        ws.column_dimensions[column].width = 22
    for column in ("B", "E", "H", "K"):
        ws.column_dimensions[column].width = 32
    ws.row_dimensions[1].height = 30

    ws["A11"] = "Notas de catálogo"
    style_section(ws["A11"], ORANGE)
    ws.merge_cells("A11:K11")
    notes = [
        "Usa los códigos exactamente como aparecen; no traduzcas AGENCY, ACTIVE, GOLD, etc.",
        "Para clientes nuevos se recomienda estado ACTIVE o INACTIVE.",
        "WARNING y BLOCKED normalmente se recalculan desde la cartera Karing.",
        "Si categoría, estado o día de pagos quedan vacíos, el importador debe aplicar los valores predeterminados.",
    ]
    for row, note in enumerate(notes, start=12):
        ws.cell(row, 1, f"• {note}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 1).font = Font(name="Aptos", size=10, color=TEXT)
        ws.row_dimensions[row].height = 24


def build_instructions(ws):
    ws.sheet_properties.tabColor = ORANGE
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A25"
    ws.merge_cells("A1:H2")
    ws["A1"] = "Plantilla de cargue masivo de clientes"
    style_title(ws["A1"], 22)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:H3")
    ws["A3"] = (
        f"Planetour CRM · Versión de plantilla 1.0 · Generada el {date.today().isoformat()} · "
        "Una fila representa un cliente."
    )
    ws["A3"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws["A3"].font = Font(name="Aptos", size=10, italic=True, color=MUTED)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 28

    ws.merge_cells("A5:H5")
    ws["A5"] = "Cómo diligenciar y entregar"
    style_section(ws["A5"], INDIGO)
    steps = [
        "Trabaja únicamente en la hoja Clientes; no cambies los nombres de sus columnas.",
        "Completa una fila por cliente. No uses celdas combinadas, fórmulas ni filas de subtítulos.",
        "Los encabezados rojos son obligatorios para la API; los naranjas son obligatorios para la operación.",
        "Elige tipo, categoría, estado y día de pagos desde las listas desplegables.",
        "Mantén NIT, IATA, documentos y teléfonos como texto para conservar ceros, signos y prefijos.",
        "Guarda el archivo como XLSX. Revisa celdas rojas/naranjas y NIT resaltados antes del cargue.",
    ]
    for row, step in enumerate(steps, start=6):
        ws.cell(row, 1, row - 5)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=INDIGO)
        ws.cell(row, 1).font = Font(name="Aptos", bold=True, color=WHITE)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 2, step)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row, 2).font = Font(name="Aptos", size=10, color=TEXT)
        ws.row_dimensions[row].height = 27

    ws.merge_cells("A13:H13")
    ws["A13"] = "Campos obligatorios"
    style_section(ws["A13"], RED)
    mandatory_headers = ["Columna", "Obligación", "Qué debe contener", "Regla"]
    for column, value in enumerate(mandatory_headers, start=1):
        cell = ws.cell(14, column, value)
        cell.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.border = THIN_BORDER
    mandatory_rows = [
        ("nombre_cliente", "API", "Razón social o nombre comercial", "2 a 200 caracteres"),
        ("tipo_cliente", "API", "AGENCY, GOVERNMENT o CORPORATE", "Debe venir del catálogo"),
        ("nit", "API", "NIT/RUT único", "3 a 50 caracteres; usar un formato consistente"),
        ("dueno_nombre", "Operativa", "Dueño o representante legal", "Obligatorio en el formulario"),
        ("pagos_nombre", "Operativa", "Encargado de pagos o cartera", "Obligatorio en el formulario"),
    ]
    for row, values in enumerate(mandatory_rows, start=15):
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row, column, value)
            cell.border = THIN_BORDER
            cell.fill = PatternFill("solid", fgColor=WHITE if row % 2 else PALE_GRAY)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 2).font = Font(
            name="Aptos",
            bold=True,
            color=RED if values[1] == "API" else ORANGE,
        )

    ws.merge_cells("F14:H14")
    ws["F14"] = "Valores predeterminados si la celda queda vacía"
    ws["F14"].fill = PatternFill("solid", fgColor=GREEN)
    ws["F14"].font = Font(name="Aptos", bold=True, color=WHITE)
    ws["F14"].alignment = Alignment(wrap_text=True)
    default_rows = [
        "categoria → GOLD",
        "estado → ACTIVE",
        "cupo_credito_cop → 0",
        "pagos_cargo → Tesorero / Cartera",
        "pagos_dias → Viernes",
    ]
    for row, value in enumerate(default_rows, start=15):
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        ws.cell(row, 6, value)
        ws.cell(row, 6).fill = PatternFill("solid", fgColor=PALE_GREEN)
        ws.cell(row, 6).border = THIN_BORDER

    ws.merge_cells("A21:H21")
    ws["A21"] = "Privacidad y campos que no se cargan"
    style_section(ws["A21"], ORANGE)
    ws.merge_cells("A22:H23")
    ws["A22"] = (
        "Los nombres, correos y teléfonos del dueño, pagos y contacto operativo pueden aparecer en la consulta "
        "interna sin autenticación; carga únicamente datos autorizados para soporte. No incluyas id, saldo Karing, "
        "días de mora, fecha de creación, contratos ni firmas GDS: la aplicación los genera, calcula o gestiona por separado."
    )
    ws["A22"].fill = PatternFill("solid", fgColor=PALE_ORANGE)
    ws["A22"].font = Font(name="Aptos", size=10, color=TEXT)
    ws["A22"].alignment = Alignment(wrap_text=True, vertical="top")

    dictionary_row = 25
    dictionary_headers = [
        "Columna Excel",
        "Nombre visible",
        "Obligatorio",
        "Tipo",
        "Límite / catálogo",
        "Predeterminado",
        "Mapeo API / JSON",
        "Descripción",
    ]
    for column, value in enumerate(dictionary_headers, start=1):
        cell = ws.cell(dictionary_row, column, value)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN_BORDER
    for row, field in enumerate(FIELDS, start=dictionary_row + 1):
        values = [
            field["key"],
            field["label"],
            "Sí — API" if field["required"] == "API" else "Sí — Operativo" if field["required"] == "Operativo" else "No",
            field["type"],
            field["limit"],
            field["default"],
            field["api"],
            field["description"],
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row, column, value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill("solid", fgColor=WHITE if row % 2 else PALE_GRAY)
        if field["required"] == "API":
            ws.cell(row, 3).font = Font(name="Aptos", bold=True, color=RED)
        elif field["required"] == "Operativo":
            ws.cell(row, 3).font = Font(name="Aptos", bold=True, color=ORANGE)
        ws.row_dimensions[row].height = 38

    add_table(
        ws,
        f"A{dictionary_row}:H{dictionary_row + len(FIELDS)}",
        "DiccionarioCamposClientes",
        "TableStyleMedium2",
    )

    widths = [22, 32, 19, 15, 34, 24, 28, 48]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_example(ws):
    ws.sheet_properties.tabColor = SKY
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 48
    for index, field in enumerate(FIELDS, start=1):
        header = ws.cell(1, index, field["key"])
        header.fill = PatternFill(
            "solid",
            fgColor=RED if field["required"] == "API" else ORANGE if field["required"] == "Operativo" else NAVY_LIGHT,
        )
        header.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(index)].width = field["width"]
        value_cell = ws.cell(2, index, field["example"])
        value_cell.border = THIN_BORDER
        value_cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)
        if field["type"] in {"Texto", "Correo"}:
            value_cell.number_format = "@"
    ws.row_dimensions[2].height = 36
    ws["A4"] = "EJEMPLO FICTICIO: no copies esta fila si contiene datos que no correspondan a un cliente real."
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(FIELDS))
    ws["A4"].fill = PatternFill("solid", fgColor=PALE_ORANGE)
    ws["A4"].font = Font(name="Aptos", bold=True, color=RED)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    add_table(ws, f"A1:{get_column_letter(len(FIELDS))}2", "EjemploCliente", "TableStyleMedium2")


def build_workbook(output_path=OUTPUT_PATH):
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "Plantilla de cargue masivo de clientes - Planetour CRM"
    workbook.properties.subject = "Datos principales y contactos de clientes"
    workbook.properties.creator = "Planetour CRM"
    workbook.properties.keywords = "Planetour, CRM, clientes, cargue masivo"
    workbook.properties.description = (
        "Plantilla XLSX con campos, catálogos, validaciones e instrucciones para preparar clientes."
    )

    clients = workbook.create_sheet("Clientes")
    instructions = workbook.create_sheet("Instrucciones")
    example = workbook.create_sheet("Ejemplo")
    catalogs = workbook.create_sheet("Catalogos")

    build_catalogs(catalogs, workbook)
    configure_client_sheet(clients)
    build_instructions(instructions)
    build_example(example)

    workbook.active = workbook.sheetnames.index("Instrucciones")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    # Abrir nuevamente detecta archivos XLSX dañados antes de entregarlos.
    verified = load_workbook(output_path, data_only=False)
    expected_sheets = ["Clientes", "Instrucciones", "Ejemplo", "Catalogos"]
    if verified.sheetnames != expected_sheets:
        raise RuntimeError(f"Hojas inesperadas: {verified.sheetnames}")
    if [verified["Clientes"].cell(1, index).value for index in range(1, len(FIELDS) + 1)] != [
        field["key"] for field in FIELDS
    ]:
        raise RuntimeError("Los encabezados de la hoja Clientes no coinciden con el modelo.")
    verified.close()
    return output_path


if __name__ == "__main__":
    result = build_workbook()
    print(result)
